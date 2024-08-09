import openpyxl
import os

def workbook_init():
    """創建一個紀錄表 history.xlsx"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "餘額紀錄表"
    sheet['A1'] = '時間'
    sheet['B1'] = '餘額'
    workbook.save("history.xlsx")
    return workbook

def check_workbook_format(workbook:openpyxl.Workbook):
    """檢查紀錄表的格式"""
    sheet = workbook.active
    return sheet['A1'].value == '時間' and sheet['B1'].value == '餘額'

def workbook_save(t, balance):
    """將今日的餘額記錄到 Excel, t : 時間, balance : 餘額"""
    if os.path.exists("history.xlsx"):
        workbook = openpyxl.load_workbook('history.xlsx')
        if not check_workbook_format(workbook):
            print("Excel 格式錯誤，重新創建紀錄表")
            os.remove("history.xlsx")
            workbook = workbook_init()
    else:
        print("創建紀錄表")
        workbook = workbook_init()
    return write_today(workbook=workbook, t=t, balance=balance)

def write_today(workbook:openpyxl.Workbook, t, balance):
    """t : 時間, balance : 餘額"""

    sheet = workbook.active
    last_row = sheet.max_row

    # 紀錄今天餘額
    sheet.cell(row=last_row + 1, column=1, value=t)
    sheet.cell(row=last_row + 1, column=2, value=balance)
    workbook.save("history.xlsx")

    # 第一筆新紀錄
    if last_row == 1:
        workbook.close()
        return "start", None, None

    # 上一個紀錄時間
    t_before = sheet.cell(row=last_row, column=1).value
    # 上一次的餘額
    balance_before = eval(sheet.cell(row=last_row, column=2).value)
    
    workbook.close()

    cost = balance_before - eval(balance)
    # 儲值了
    if cost < 0:
        return "top_up", t_before, balance_before
    # 返回距離上次花了多少
    else:
        return "success", t_before, cost
